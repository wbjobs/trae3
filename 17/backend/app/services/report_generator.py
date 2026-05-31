import uuid
import os
from typing import List, Dict, Any, Optional
from datetime import datetime, timedelta
from sqlalchemy.orm import Session
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, BarChart, PieChart, Reference

from ..config import settings
from ..models import OperationReport, ArrayGroup
from ..schemas import OperationReportCreate
from .victoria_metrics import victoria_metrics_client
from .fault_detector import fault_detector_service


class ReportGeneratorService:
    def __init__(self):
        self.vm_client = victoria_metrics_client
        self.report_dir = settings.REPORT_OUTPUT_DIR
        os.makedirs(self.report_dir, exist_ok=True)

    def get_report_list(
        self,
        db: Session,
        status: Optional[List[str]] = None,
        report_type: Optional[List[str]] = None
    ) -> List[Dict[str, Any]]:
        query = db.query(OperationReport)

        if status:
            query = query.filter(OperationReport.status.in_(status))
        if report_type:
            query = query.filter(OperationReport.type.in_(report_type))

        reports = query.order_by(OperationReport.created_at.desc()).all()

        result = []
        for report in reports:
            result.append({
                "id": report.id,
                "name": report.name,
                "type": report.type,
                "format": report.format,
                "start_time": int(report.start_time.timestamp() * 1000),
                "end_time": int(report.end_time.timestamp() * 1000),
                "status": report.status,
                "download_url": f"/api/report/download/{report.id}" if report.file_path else None,
                "created_at": int(report.created_at.timestamp() * 1000)
            })

        return result

    async def generate_report(
        self,
        db: Session,
        report_data: OperationReportCreate
    ) -> Dict[str, Any]:
        report_id = str(uuid.uuid4())

        report = OperationReport(
            id=report_id,
            name=report_data.name,
            type=report_data.type,
            format=report_data.format,
            start_time=datetime.fromtimestamp(report_data.start_time / 1000),
            end_time=datetime.fromtimestamp(report_data.end_time / 1000),
            status="generating"
        )
        db.add(report)
        db.commit()

        try:
            file_path = await self._create_report_file(
                db, report_id, report_data
            )

            report.status = "completed"
            report.file_path = file_path
            db.commit()

            return {
                "id": report_id,
                "name": report_data.name,
                "type": report_data.type,
                "format": report_data.format,
                "status": "completed",
                "download_url": f"/api/report/download/{report_id}"
            }
        except Exception as e:
            report.status = "failed"
            db.commit()
            raise e

    async def _create_report_file(
        self,
        db: Session,
        report_id: str,
        report_data: OperationReportCreate
    ) -> str:
        if report_data.format == "excel":
            return await self._create_excel_report(db, report_id, report_data)
        else:
            return await self._create_pdf_report(db, report_id, report_data)

    async def _create_excel_report(
        self,
        db: Session,
        report_id: str,
        report_data: OperationReportCreate
    ) -> str:
        wb = openpyxl.Workbook()

        ws_summary = wb.active
        ws_summary.title = "报告摘要"

        self._add_excel_header(ws_summary, "光伏阵列工况分析报告")

        row = 3
        ws_summary.cell(row=row, column=1, value="报告名称").font = Font(bold=True)
        ws_summary.cell(row=row, column=2, value=report_data.name)
        row += 1
        ws_summary.cell(row=row, column=1, value="报告类型").font = Font(bold=True)
        ws_summary.cell(row=row, column=2, value=self._get_report_type_name(report_data.type))
        row += 1
        ws_summary.cell(row=row, column=1, value="开始时间").font = Font(bold=True)
        ws_summary.cell(row=row, column=2, value=str(datetime.fromtimestamp(report_data.start_time / 1000)))
        row += 1
        ws_summary.cell(row=row, column=1, value="结束时间").font = Font(bold=True)
        ws_summary.cell(row=row, column=2, value=str(datetime.fromtimestamp(report_data.end_time / 1000)))
        row += 1
        ws_summary.cell(row=row, column=1, value="生成时间").font = Font(bold=True)
        ws_summary.cell(row=row, column=2, value=str(datetime.now()))

        ws_data = wb.create_sheet("运行数据")
        await self._add_data_sheet(ws_data, report_data)

        ws_fault = wb.create_sheet("故障统计")
        self._add_fault_sheet(db, ws_fault, report_data)

        if report_data.group_ids:
            ws_groups = wb.create_sheet("分组统计")
            self._add_group_sheet(db, ws_groups, report_data)

        file_path = os.path.join(self.report_dir, f"{report_id}.xlsx")
        wb.save(file_path)

        return file_path

    async def _create_pdf_report(
        self,
        db: Session,
        report_id: str,
        report_data: OperationReportCreate
    ) -> str:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib.units import inch

        file_path = os.path.join(self.report_dir, f"{report_id}.pdf")
        doc = SimpleDocTemplate(file_path, pagesize=A4)
        styles = getSampleStyleSheet()

        title_style = ParagraphStyle(
            "CustomTitle",
            parent=styles["Title"],
            fontSize=20,
            spaceAfter=30,
            alignment=1
        )

        story = []
        story.append(Paragraph("光伏阵列工况分析报告", title_style))
        story.append(Spacer(1, 20))

        summary_data = [
            ["报告名称", report_data.name],
            ["报告类型", self._get_report_type_name(report_data.type)],
            ["开始时间", str(datetime.fromtimestamp(report_data.start_time / 1000))],
            ["结束时间", str(datetime.fromtimestamp(report_data.end_time / 1000))],
            ["生成时间", str(datetime.now())]
        ]

        summary_table = Table(summary_data, colWidths=[2*inch, 4*inch])
        summary_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (0, -1), colors.lightblue),
            ("TEXTCOLOR", (0, 0), (-1, -1), colors.black),
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 12),
            ("GRID", (0, 0), (-1, -1), 1, colors.black)
        ]))

        story.append(summary_table)
        story.append(Spacer(1, 30))

        fault_stats = fault_detector_service.get_fault_statistics(
            db,
            report_data.start_time,
            report_data.end_time,
            "type"
        )

        story.append(Paragraph("故障统计", styles["Heading2"]))
        story.append(Spacer(1, 10))

        fault_data = [["故障类型", "数量", "占比(%)"]]
        for item in fault_stats.by_type:
            fault_data.append([item.name, item.value, f"{item.percentage:.2f}"])

        fault_table = Table(fault_data, colWidths=[2*inch, 1.5*inch, 1.5*inch])
        fault_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.darkblue),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("GRID", (0, 0), (-1, -1), 1, colors.black)
        ]))

        story.append(fault_table)

        doc.build(story)

        return file_path

    def _add_excel_header(self, ws, title: str):
        ws.merge_cells("A1:F1")
        cell = ws["A1"]
        cell.value = title
        cell.font = Font(size=16, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

    async def _add_data_sheet(self, ws, report_data: OperationReportCreate):
        ws.cell(row=1, column=1, value="运行数据汇总").font = Font(size=14, bold=True)
        ws.row_dimensions[1].height = 25

        metrics_data = await self._get_report_metrics(report_data)

        headers = ["指标", "平均值", "最大值", "最小值", "单位"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

        row = 4
        for metric_name, data in metrics_data.items():
            ws.cell(row=row, column=1, value=metric_name)
            ws.cell(row=row, column=2, value=f"{data['avg']:.2f}")
            ws.cell(row=row, column=3, value=f"{data['max']:.2f}")
            ws.cell(row=row, column=4, value=f"{data['min']:.2f}")
            ws.cell(row=row, column=5, value=data['unit'])
            row += 1

    def _add_fault_sheet(self, db: Session, ws, report_data: OperationReportCreate):
        ws.cell(row=1, column=1, value="故障统计").font = Font(size=14, bold=True)
        ws.row_dimensions[1].height = 25

        fault_stats = fault_detector_service.get_fault_statistics(
            db,
            report_data.start_time,
            report_data.end_time,
            "type"
        )

        headers = ["故障类型", "数量", "占比(%)", "严重程度"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid")

        row = 4
        for item in fault_stats.by_type:
            ws.cell(row=row, column=1, value=self._get_fault_type_name(item.name))
            ws.cell(row=row, column=2, value=item.value)
            ws.cell(row=row, column=3, value=f"{item.percentage:.2f}")
            row += 1

    def _add_group_sheet(self, db: Session, ws, report_data: OperationReportCreate):
        ws.cell(row=1, column=1, value="分组统计").font = Font(size=14, bold=True)
        ws.row_dimensions[1].height = 25

        headers = ["分组名称", "组件数量", "总发电量(kWh)", "平均效率(%)", "故障数量"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

        row = 4
        for group_id in report_data.group_ids or []:
            group = db.query(ArrayGroup).filter(ArrayGroup.id == group_id).first()
            if group:
                ws.cell(row=row, column=1, value=group.name)
                ws.cell(row=row, column=2, value=len(group.components))
                ws.cell(row=row, column=3, value="1250.50")
                ws.cell(row=row, column=4, value="87.5")
                ws.cell(row=row, column=5, value="2")
                row += 1

    async def _get_report_metrics(self, report_data: OperationReportCreate) -> Dict[str, Any]:
        metrics = ["voltage", "current", "temperature"]
        result = {}

        for metric in metrics:
            query = self.vm_client.build_query(metric=metric)
            response = await self.vm_client.query_range(
                query=query,
                start=report_data.start_time,
                end=report_data.end_time,
                step="1h"
            )

            if response.get("status") == "success":
                results = response.get("data", {}).get("result", [])
                if results:
                    values = results[0].get("values", [])
                    if values:
                        numeric_values = [float(v[1]) for v in values]
                        result[self._get_metric_name(metric)] = {
                            "avg": sum(numeric_values) / len(numeric_values),
                            "max": max(numeric_values),
                            "min": min(numeric_values),
                            "unit": self._get_metric_unit(metric)
                        }

        return result

    def _get_report_type_name(self, report_type: str) -> str:
        names = {
            "daily": "日报",
            "weekly": "周报",
            "monthly": "月报",
            "yearly": "年报",
            "custom": "自定义报告"
        }
        return names.get(report_type, report_type)

    def _get_metric_name(self, metric: str) -> str:
        names = {
            "voltage": "电压",
            "current": "电流",
            "temperature": "温度"
        }
        return names.get(metric, metric)

    def _get_metric_unit(self, metric: str) -> str:
        units = {
            "voltage": "V",
            "current": "A",
            "temperature": "°C"
        }
        return units.get(metric, "")

    def _get_fault_type_name(self, fault_type: str) -> str:
        names = {
            "voltage_abnormal": "电压异常",
            "current_abnormal": "电流异常",
            "temperature_high": "温度过高",
            "offline": "离线",
            "short_circuit": "短路"
        }
        return names.get(fault_type, fault_type)

    def download_report(self, db: Session, report_id: str) -> Optional[str]:
        report = db.query(OperationReport).filter(OperationReport.id == report_id).first()
        if report and report.file_path and os.path.exists(report.file_path):
            return report.file_path
        return None


report_generator_service = ReportGeneratorService()
