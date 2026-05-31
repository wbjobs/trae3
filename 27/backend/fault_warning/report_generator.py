"""运维报表生成器 - 修复统计计算错误"""
from datetime import datetime, timedelta, date
from typing import Optional
import os
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.conf import settings

from api.models import MaintenanceReport, Device, FaultRecord
from api.influxdb_client import influx_client


class ReportGenerator:
    """运维报表生成器 - 修正版"""
    
    def __init__(self):
        self.report_dir = os.path.join(settings.BASE_DIR, 'reports')
        os.makedirs(self.report_dir, exist_ok=True)
    
    def generate_daily_report(self, report_date: Optional[date] = None) -> Optional[MaintenanceReport]:
        """生成日报表 - 修正统计逻辑"""
        if not report_date:
            report_date = datetime.now().date()
        
        start_time = datetime.combine(report_date, datetime.min.time())
        end_time = start_time + timedelta(days=1)
        
        total_devices = Device.objects.count()
        online_devices = Device.objects.filter(status='online').count()
        
        faults_today = FaultRecord.objects.filter(
            fault_time__gte=start_time,
            fault_time__lt=end_time
        )
        total_faults = faults_today.count()
        resolved_faults = faults_today.filter(resolved=True).count()
        
        fault_device_ids = faults_today.values_list('device_id', flat=True).distinct()
        fault_devices = len(set(fault_device_ids))
        
        online_rate = round(online_devices / total_devices * 100, 2) if total_devices > 0 else 0
        fault_rate = round(fault_devices / total_devices * 100, 2) if total_devices > 0 else 0
        resolution_rate = round(resolved_faults / total_faults * 100, 2) if total_faults > 0 else 0
        
        pressure_data = influx_client.query_pressure_data(
            start_time=start_time,
            end_time=end_time,
            aggregation='1h'
        )
        flow_data = influx_client.query_flow_data(
            start_time=start_time,
            end_time=end_time,
            aggregation='1h'
        )
        
        df_p = pd.DataFrame(pressure_data) if pressure_data else pd.DataFrame()
        df_f = pd.DataFrame(flow_data) if flow_data else pd.DataFrame()
        
        if not df_p.empty and 'value' in df_p.columns and 'pressure' not in df_p.columns:
            df_p['pressure'] = df_p['value']
        if not df_f.empty and 'value' in df_f.columns and 'flow' not in df_f.columns:
            df_f['flow'] = df_f['value']
        
        avg_pressure = 0.45
        if not df_p.empty:
            pressure_col = 'pressure' if 'pressure' in df_p.columns else 'value'
            if pressure_col in df_p.columns:
                avg_pressure = float(df_p[pressure_col].mean()) if len(df_p) > 0 else 0.45
        
        avg_flow = 100.0
        if not df_f.empty:
            flow_col = 'flow' if 'flow' in df_f.columns else 'value'
            if flow_col in df_f.columns:
                avg_flow = float(df_f[flow_col].mean()) if len(df_f) > 0 else 100.0
        
        report, created = MaintenanceReport.objects.update_or_create(
            report_date=report_date,
            defaults={
                'total_devices': total_devices,
                'online_devices': online_devices,
                'fault_devices': fault_devices,
                'total_faults': total_faults,
                'resolved_faults': resolved_faults,
                'avg_pressure': round(avg_pressure, 3),
                'avg_flow': round(avg_flow, 2),
                'online_rate': online_rate,
                'fault_rate': fault_rate,
                'resolution_rate': resolution_rate,
                'report_notes': f"自动生成日报表 - 在线率: {online_rate}%, 故障率: {fault_rate}%, 解决率: {resolution_rate}%"
            }
        )
        
        try:
            report_file = self._generate_excel_report(report, df_p, df_f, faults_today)
            report.report_file = report_file
            report.save()
        except Exception as e:
            print(f"生成Excel报表失败: {e}")
        
        return report
    
    def _generate_excel_report(self, report: MaintenanceReport, 
                                df_p: pd.DataFrame, 
                                df_f: pd.DataFrame,
                                faults) -> str:
        """生成Excel报表"""
        filename = f"report_{report.report_date.strftime('%Y%m%d')}.xlsx"
        filepath = os.path.join(self.report_dir, filename)
        
        wb = Workbook()
        
        self._create_summary_sheet(wb, report)
        self._create_pressure_sheet(wb, df_p)
        self._create_flow_sheet(wb, df_f)
        self._create_fault_sheet(wb, faults)
        
        wb.save(filepath)
        return filepath
    
    def _create_summary_sheet(self, wb: Workbook, report: MaintenanceReport):
        """创建概览表 - 修正计算逻辑"""
        ws = wb.active
        ws.title = "运维概览"
        
        header_font = Font(bold=True, size=14, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center")
        
        ws['A1'] = "供水管网运维日报表"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:F1')
        ws['A1'].alignment = center_align
        
        ws['A2'] = f"报表日期: {report.report_date.strftime('%Y年%m月%d日')}"
        ws.merge_cells('A2:F2')
        ws['A2'].alignment = center_align
        
        headers = ["指标名称", "数值", "指标名称", "数值", "指标名称", "数值"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
        
        online_rate_value = getattr(report, 'online_rate', 
            round(report.online_devices / report.total_devices * 100, 1) if report.total_devices > 0 else 0)
        fault_rate_value = getattr(report, 'fault_rate',
            round(report.fault_devices / report.total_devices * 100, 1) if report.total_devices > 0 else 0)
        resolution_rate_value = getattr(report, 'resolution_rate',
            round(report.resolved_faults / report.total_faults * 100, 1) if report.total_faults > 0 else 0)
        
        data = [
            ["设备总数", report.total_devices, "在线设备", report.online_devices, "在线率", f"{online_rate_value:.1f}%"],
            ["故障设备数", report.fault_devices, "故障率", f"{fault_rate_value:.1f}%", "设备完好率", f"{max(0, 100 - fault_rate_value):.1f}%"],
            ["今日故障数", report.total_faults, "已解决", report.resolved_faults, "解决率", f"{resolution_rate_value:.1f}%"],
            ["平均压力", f"{report.avg_pressure:.3f} MPa", "平均流量", f"{report.avg_flow:.2f} m³/h", "", ""]
        ]
        
        for row_idx, row_data in enumerate(data, 5):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = center_align
        
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws.column_dimensions[col].width = 18
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in range(4, 9):
            for col in range(1, 7):
                ws.cell(row=row, column=col).border = thin_border
    
    def _create_pressure_sheet(self, wb: Workbook, df: pd.DataFrame):
        """创建压力数据表 - 修复空数据处理"""
        ws = wb.create_sheet("压力统计")
        
        if df.empty:
            ws['A1'] = "今日暂无压力数据"
            ws['A1'].font = Font(color="888888")
            return
        
        pressure_col = 'pressure' if 'pressure' in df.columns else 'value'
        if pressure_col not in df.columns:
            ws['A1'] = "压力数据格式错误"
            return
        
        zones = df['zone'].unique() if 'zone' in df.columns else ['全部']
        zone_data = []
        
        for zone in zones:
            if zone == '全部':
                zone_df = df
            else:
                zone_df = df[df['zone'] == zone]
            
            if len(zone_df) == 0:
                continue
            
            values = zone_df[pressure_col].dropna()
            
            zone_data.append({
                '分区': zone,
                '平均值(MPa)': round(float(values.mean()), 3) if len(values) > 0 else 0,
                '最小值(MPa)': round(float(values.min()), 3) if len(values) > 0 else 0,
                '最大值(MPa)': round(float(values.max()), 3) if len(values) > 0 else 0,
                '标准差': round(float(values.std()), 3) if len(values) > 1 else 0,
                '数据点数': len(values),
                '完好率(%)': self._calculate_completeness(len(values), 24)
            })
        
        result_df = pd.DataFrame(zone_data)
        
        if len(result_df) == 0:
            ws['A1'] = "无有效压力数据"
            return
        
        headers = result_df.columns.tolist()
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        
        for row_idx, row_data in enumerate(result_df.values, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col)].width = 16
    
    def _create_flow_sheet(self, wb: Workbook, df: pd.DataFrame):
        """创建流量数据表 - 修复空数据处理"""
        ws = wb.create_sheet("流量统计")
        
        if df.empty:
            ws['A1'] = "今日暂无流量数据"
            ws['A1'].font = Font(color="888888")
            return
        
        flow_col = 'flow' if 'flow' in df.columns else 'value'
        if flow_col not in df.columns:
            ws['A1'] = "流量数据格式错误"
            return
        
        zones = df['zone'].unique() if 'zone' in df.columns else ['全部']
        zone_data = []
        
        for zone in zones:
            if zone == '全部':
                zone_df = df
            else:
                zone_df = df[df['zone'] == zone]
            
            if len(zone_df) == 0:
                continue
            
            values = zone_df[flow_col].dropna()
            
            zone_data.append({
                '分区': zone,
                '平均值(m³/h)': round(float(values.mean()), 2) if len(values) > 0 else 0,
                '最小值(m³/h)': round(float(values.min()), 2) if len(values) > 0 else 0,
                '最大值(m³/h)': round(float(values.max()), 2) if len(values) > 0 else 0,
                '总流量(m³)': round(float(values.sum()) * 0.25, 2) if len(values) > 0 else 0,
                '数据点数': len(values),
                '完好率(%)': self._calculate_completeness(len(values), 24)
            })
        
        result_df = pd.DataFrame(zone_data)
        
        if len(result_df) == 0:
            ws['A1'] = "无有效流量数据"
            return
        
        headers = result_df.columns.tolist()
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        
        for row_idx, row_data in enumerate(result_df.values, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col)].width = 18
    
    def _create_fault_sheet(self, wb: Workbook, faults):
        """创建故障记录表 - 修复空数据处理"""
        ws = wb.create_sheet("故障记录")
        
        fault_type_names = {
            'low_pressure': '压力过低',
            'high_pressure': '压力过高',
            'low_flow': '流量过低',
            'high_flow': '流量过高',
            'offline': '设备离线',
            'abnormal': '数据异常'
        }
        
        headers = ["序号", "设备ID", "设备名称", "分区", "故障类型", "故障时间", "故障值", "阈值", "状态", "解决时间"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        
        fault_list = list(faults)
        
        if len(fault_list) == 0:
            ws.cell(row=2, column=1, value="今日无故障记录")
            ws.cell(row=2, column=1).font = Font(color="00B050")
        else:
            for row_idx, fault in enumerate(fault_list, 2):
                device_id = getattr(fault.device, 'device_id', '') if fault.device else ''
                device_name = getattr(fault.device, 'name', '') if fault.device else ''
                zone_name = getattr(fault.zone, 'name', '') if fault.zone else ''
                
                ws.cell(row=row_idx, column=1, value=row_idx - 1)
                ws.cell(row=row_idx, column=2, value=device_id)
                ws.cell(row=row_idx, column=3, value=device_name)
                ws.cell(row=row_idx, column=4, value=zone_name)
                ws.cell(row=row_idx, column=5, value=fault_type_names.get(fault.fault_type, fault.fault_type))
                ws.cell(row=row_idx, column=6, value=fault.fault_time.strftime('%Y-%m-%d %H:%M') if fault.fault_time else '')
                ws.cell(row=row_idx, column=7, value=round(fault.fault_value, 3) if fault.fault_value else 0)
                ws.cell(row=row_idx, column=8, value=fault.threshold)
                ws.cell(row=row_idx, column=9, value="已解决" if fault.resolved else "待处理")
                ws.cell(row=row_idx, column=10, value=fault.resolved_time.strftime('%Y-%m-%d %H:%M') if fault.resolved_time else '')
        
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
            ws.column_dimensions[col].width = 16
    
    def _calculate_completeness(self, actual_count: int, expected_count: int) -> float:
        """计算数据完整率"""
        if expected_count == 0:
            return 100.0
        return round(min(100.0, actual_count / expected_count * 100), 1)
