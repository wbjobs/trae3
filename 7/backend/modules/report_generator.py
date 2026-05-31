import sys
import os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from typing import List, Dict, Optional
from datetime import datetime, timedelta
import pandas as pd
import numpy as np
import logging
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from config import Config

logger = logging.getLogger(__name__)

METRIC_THRESHOLDS = Config.METRIC_CONFIGS

class ReportGeneratorService:
    def __init__(self):
        Config.ensure_dirs()
        self.report_dir = Config.REPORT_OUTPUT_DIR
        self.export_dir = Config.EXPORT_DIR

    def generate_maintenance_report(self, device_id: str, time_range: str,
                                     statistics: Dict, format: str = 'pdf') -> Dict:
        try:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"maintenance_report_{device_id}_{timestamp}.{format}"
            filepath = os.path.join(self.report_dir, filename)

            if format == 'pdf':
                self._generate_pdf_report(filepath, device_id, time_range, statistics)
            elif format == 'xlsx':
                self._generate_excel_report(filepath, device_id, time_range, statistics)
            else:
                return {'success': False, 'message': f'Unsupported format: {format}'}

            return {
                'success': True,
                'filename': filename,
                'filepath': filepath,
                'download_url': f'/api/reports/download/{filename}'
            }
        except Exception as e:
            logger.error(f"Error generating maintenance report: {e}")
            return {'success': False, 'message': str(e)}

    def _generate_pdf_report(self, filepath: str, device_id: str,
                              time_range: str, statistics: Dict):
        doc = SimpleDocTemplate(filepath, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []

        title_style = ParagraphStyle('CustomTitle', parent=styles['Title'],
            fontSize=24, spaceAfter=30, textColor=colors.HexColor('#1a365d'))
        heading_style = ParagraphStyle('CustomHeading', parent=styles['Heading2'],
            fontSize=16, spaceAfter=15, textColor=colors.HexColor('#2b6cb0'))
        normal_style = styles['Normal']

        story.append(Paragraph(f"设备运维报告 - {device_id}", title_style))
        story.append(Paragraph(f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", normal_style))
        story.append(Paragraph(f"时间范围: {time_range}", normal_style))
        story.append(Spacer(1, 0.5 * inch))

        story.append(Paragraph("一、设备运行统计概览", heading_style))

        overview_data = [['指标', '最小值', '最大值', '平均值', '中位数', '标准偏差', 'P95', 'P99', '最新值']]
        for metric, stats in statistics.get('statistics', {}).items():
            overview_data.append([
                metric,
                str(stats.get('min', '-')),
                str(stats.get('max', '-')),
                str(stats.get('avg', '-')),
                str(stats.get('median', stats.get('avg', '-'))),
                str(stats.get('std', '-')),
                str(stats.get('p95', '-')),
                str(stats.get('p99', '-')),
                str(stats.get('latest', '-'))
            ])

        if len(overview_data) > 1:
            table = Table(overview_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2b6cb0')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#ebf8ff')),
                ('GRID', (0, 0), (-1, -1), 1, colors.gray)
            ]))
            story.append(table)
        else:
            story.append(Paragraph("暂无统计数据", normal_style))

        story.append(Spacer(1, 0.3 * inch))
        story.append(Paragraph("二、运维建议", heading_style))

        recommendations = self._generate_recommendations(statistics)
        for rec in recommendations:
            story.append(Paragraph(f"• {rec}", normal_style))
            story.append(Spacer(1, 0.1 * inch))

        story.append(Spacer(1, 0.3 * inch))
        story.append(Paragraph("三、报告说明", heading_style))
        story.append(Paragraph("本报告由工业设备运维时序数据分析平台自动生成。", normal_style))
        story.append(Paragraph("数据来源于设备实时监控系统，用于设备状态评估和维护决策。", normal_style))

        doc.build(story)
        logger.info(f"PDF report generated: {filepath}")

    def _generate_excel_report(self, filepath: str, device_id: str,
                                time_range: str, statistics: Dict):
        wb = Workbook()
        ws = wb.active
        ws.title = "运维报告"

        header_font = Font(name='微软雅黑', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='2B6CB0', end_color='2B6CB0', fill_type='solid')
        title_font = Font(name='微软雅黑', size=12, bold=True)
        normal_font = Font(name='微软雅黑', size=11)
        center_align = Alignment(horizontal='center', vertical='center')

        ws['A1'] = f"设备运维报告 - {device_id}"
        ws['A1'].font = Font(name='微软雅黑', size=20, bold=True, color='1A365D')
        ws.merge_cells('A1:I1')

        ws['A2'] = f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A2'].font = normal_font
        ws.merge_cells('A2:I2')

        ws['A3'] = f"时间范围: {time_range}"
        ws['A3'].font = normal_font
        ws.merge_cells('A3:I3')

        row = 5
        ws.cell(row=row, column=1, value='一、设备运行统计概览').font = title_font
        ws.merge_cells(f'A{row}:I{row}')

        row += 2
        headers = ['指标', '最小值', '最大值', '平均值', '中位数', '标准偏差', 'P95', 'P99', '最新值']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align

        row += 1
        for metric, stats in statistics.get('statistics', {}).items():
            ws.cell(row=row, column=1, value=metric).font = normal_font
            ws.cell(row=row, column=2, value=stats.get('min', 0)).font = normal_font
            ws.cell(row=row, column=3, value=stats.get('max', 0)).font = normal_font
            ws.cell(row=row, column=4, value=stats.get('avg', 0)).font = normal_font
            ws.cell(row=row, column=5, value=stats.get('median', stats.get('avg', 0))).font = normal_font
            ws.cell(row=row, column=6, value=stats.get('std', 0)).font = normal_font
            ws.cell(row=row, column=7, value=stats.get('p95', 0)).font = normal_font
            ws.cell(row=row, column=8, value=stats.get('p99', 0)).font = normal_font
            ws.cell(row=row, column=9, value=stats.get('latest', 0)).font = normal_font
            row += 1

        row += 2
        ws.cell(row=row, column=1, value='二、运维建议').font = title_font
        ws.merge_cells(f'A{row}:I{row}')

        row += 2
        recommendations = self._generate_recommendations(statistics)
        for i, rec in enumerate(recommendations):
            ws.cell(row=row + i, column=1, value=f'• {rec}').font = normal_font
            ws.merge_cells(f'A{row + i}:I{row + i}')

        for col in range(1, 10):
            ws.column_dimensions[chr(64 + col)].width = 14

        wb.save(filepath)
        logger.info(f"Excel report generated: {filepath}")

    def _generate_recommendations(self, statistics: Dict) -> List[str]:
        recommendations = []
        stats = statistics.get('statistics', {})

        for metric, metric_stats in stats.items():
            config = METRIC_THRESHOLDS.get(metric)
            if not config:
                continue

            avg = metric_stats.get('avg', 0)
            latest = metric_stats.get('latest', 0)
            std = metric_stats.get('std', 0)
            p95 = metric_stats.get('p95', 0)
            max_val = metric_stats.get('max', 0)
            warn_high = config.get('warn_high', config['max'])
            warn_low = config.get('warn_low', config['min'])
            physical_max = config['max']
            physical_min = config['min']

            if latest > warn_high:
                recommendations.append(
                    f"{metric} 最新值 {latest} 超过预警上限 {warn_high}，建议立即检查设备状态"
                )
            elif latest > warn_high * 0.85:
                recommendations.append(
                    f"{metric} 最新值 {latest} 接近预警上限 {warn_high}，建议安排维护检查"
                )

            if latest < warn_low:
                recommendations.append(
                    f"{metric} 最新值 {latest} 低于预警下限 {warn_low}，建议检查设备是否异常停机"
                )

            if max_val > physical_max:
                recommendations.append(
                    f"{metric} 最大值 {max_val} 超出物理量程上限 {physical_max}，数据可能异常"
                )

            if std > config.get('std', 0) * 2:
                recommendations.append(
                    f"{metric} 波动较大（标准差={std}），可能存在设备不稳定或传感器故障"
                )

            if avg > 0 and p95 > avg * 1.5:
                recommendations.append(
                    f"{metric} P95值 {p95} 远高于均值 {avg}，存在间歇性尖峰，建议排查"
                )

        if not recommendations:
            recommendations.append("设备运行状态良好，各项指标在正常范围内，建议继续保持常规维护巡检")

        return recommendations

    def generate_group_report(self, group_id: str, time_range: str,
                               group_stats: Dict, device_stats: Dict,
                               format: str = 'pdf') -> Dict:
        try:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"group_report_{group_id}_{timestamp}.{format}"
            filepath = os.path.join(self.report_dir, filename)

            if format == 'pdf':
                self._generate_group_pdf_report(filepath, group_id, time_range, group_stats, device_stats)
            elif format == 'xlsx':
                self._generate_group_excel_report(filepath, group_id, time_range, group_stats, device_stats)
            else:
                return {'success': False, 'message': f'Unsupported format: {format}'}

            return {
                'success': True, 'filename': filename, 'filepath': filepath,
                'download_url': f'/api/reports/download/{filename}'
            }
        except Exception as e:
            logger.error(f"Error generating group report: {e}")
            return {'success': False, 'message': str(e)}

    def _generate_group_pdf_report(self, filepath: str, group_id: str, time_range: str,
                                    group_stats: Dict, device_stats: Dict):
        doc = SimpleDocTemplate(filepath, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []

        title_style = ParagraphStyle('CustomTitle', parent=styles['Title'],
            fontSize=24, spaceAfter=30, textColor=colors.HexColor('#1a365d'))
        heading_style = ParagraphStyle('CustomHeading', parent=styles['Heading2'],
            fontSize=16, spaceAfter=15, textColor=colors.HexColor('#2b6cb0'))
        normal_style = styles['Normal']

        story.append(Paragraph(f"设备组运维报告 - {group_stats.get('group_name', group_id)}", title_style))
        story.append(Paragraph(f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", normal_style))
        story.append(Paragraph(f"时间范围: {time_range}", normal_style))
        story.append(Spacer(1, 0.3 * inch))

        story.append(Paragraph("一、设备组概览", heading_style))
        overview_data = [
            ['统计项', '数值'],
            ['设备总数', str(group_stats.get('total_devices', 0))],
            ['运行中', str(group_stats.get('status_distribution', {}).get('running', 0))],
            ['待机', str(group_stats.get('status_distribution', {}).get('standby', 0))],
            ['故障', str(group_stats.get('status_distribution', {}).get('fault', 0))]
        ]
        table = Table(overview_data, colWidths=[2*inch, 2*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2b6cb0')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.gray)
        ]))
        story.append(table)
        story.append(Spacer(1, 0.3 * inch))

        story.append(Paragraph("二、各设备统计数据", heading_style))
        device_data = [['设备ID', '状态', '平均温度', '平均振动']]
        for device_id, stats in device_stats.items():
            device_data.append([
                device_id,
                stats.get('status', 'unknown'),
                str(stats.get('avg_temp', '-')),
                str(stats.get('avg_vibration', '-'))
            ])
        if len(device_data) > 1:
            device_table = Table(device_data)
            device_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#38a169')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.gray),
                ('FONTSIZE', (0, 0), (-1, -1), 8)
            ]))
            story.append(device_table)

        doc.build(story)
        logger.info(f"Group PDF report generated: {filepath}")

    def _generate_group_excel_report(self, filepath: str, group_id: str, time_range: str,
                                      group_stats: Dict, device_stats: Dict):
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "组概览"

        ws1['A1'] = f"设备组运维报告 - {group_stats.get('group_name', group_id)}"
        ws1['A1'].font = Font(name='微软雅黑', size=16, bold=True)
        ws1.merge_cells('A1:D1')

        ws1['A3'] = "设备总数"
        ws1['B3'] = group_stats.get('total_devices', 0)
        ws1['A4'] = "运行中"
        ws1['B4'] = group_stats.get('status_distribution', {}).get('running', 0)
        ws1['A5'] = "待机"
        ws1['B5'] = group_stats.get('status_distribution', {}).get('standby', 0)
        ws1['A6'] = "故障"
        ws1['B6'] = group_stats.get('status_distribution', {}).get('fault', 0)

        ws2 = wb.create_sheet("设备详情")
        ws2.append(['设备ID', '状态', '平均温度', '平均振动'])
        for device_id, stats in device_stats.items():
            ws2.append([
                device_id,
                stats.get('status', 'unknown'),
                stats.get('avg_temp', 0),
                stats.get('avg_vibration', 0)
            ])

        wb.save(filepath)
        logger.info(f"Group Excel report generated: {filepath}")

    def export_data_to_csv(self, data: pd.DataFrame, filename_prefix: str) -> Dict:
        try:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"{filename_prefix}_{timestamp}.csv"
            filepath = os.path.join(self.export_dir, filename)
            data.to_csv(filepath, index=False, encoding='utf-8-sig')
            return {
                'success': True, 'filename': filename, 'filepath': filepath,
                'download_url': f'/api/exports/download/{filename}'
            }
        except Exception as e:
            logger.error(f"Error exporting CSV: {e}")
            return {'success': False, 'message': str(e)}

    def get_report_list(self) -> Dict:
        try:
            reports = []
            if os.path.exists(self.report_dir):
                for filename in os.listdir(self.report_dir):
                    filepath = os.path.join(self.report_dir, filename)
                    if os.path.isfile(filepath):
                        stat = os.stat(filepath)
                        reports.append({
                            'filename': filename,
                            'size': stat.st_size,
                            'created_at': datetime.fromtimestamp(stat.st_ctime).isoformat(),
                            'download_url': f'/api/reports/download/{filename}'
                        })
            reports.sort(key=lambda x: x['created_at'], reverse=True)
            return {'success': True, 'reports': reports, 'total': len(reports)}
        except Exception as e:
            logger.error(f"Error getting report list: {e}")
            return {'success': False, 'message': str(e)}

report_generator_service = ReportGeneratorService()
